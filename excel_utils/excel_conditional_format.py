#!/usr/bin/env python3
import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import IconSetRule, Rule, IconSet, FormatObject
from openpyxl.utils.cell import range_boundaries
def _parse_number_string(s: str):
    """Try to parse common numeric string formats to float.

    Handles currency symbols ($,€,£,¥), thousands separators, parentheses for negatives,
    leading/trailing spaces, and percentages. Returns (parsed_float, success_bool).
    """
    if s is None:
        return None, False
    if not isinstance(s, str):
        # Already numeric?
        try:
            return float(s), True
        except Exception:
            return None, False

    raw = s.strip()
    if raw == "":
        return None, False

    # Remove currency symbols and spaces
    for ch in "$€£¥":
        raw = raw.replace(ch, "")
    raw = raw.replace(",", "")

    is_percent = False
    if raw.endswith("%"):
        is_percent = True
        raw = raw[:-1]

    negative = False
    # Parentheses denote negative numbers, e.g., (123)
    if raw.startswith("(") and raw.endswith(")"):
        negative = True
        raw = raw[1:-1]

    try:
        val = float(raw)
        if is_percent:
            val = val / 100.0
        if negative:
            val = -val
        return val, True
    except Exception:
        return None, False


def _coerce_range_to_numbers(ws, cell_range: str, non_numeric_to_zero: bool = False):
    """Coerce cells in range to numbers when possible.

    If `non_numeric_to_zero` is True, set unparsable cells to 0, otherwise leave as-is.
    """
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            parsed, ok = _parse_number_string(cell.value)
            if ok:
                # Try to cast to int if it is integral
                if parsed.is_integer():
                    cell.value = int(parsed)
                else:
                    cell.value = parsed
            elif non_numeric_to_zero:
                cell.value = 0


ALLOWED_ICON_STYLES = [
    "3Arrows",
    "3ArrowsGray",
    "3Flags",
    "3TrafficLights1",
    "3TrafficLights2",
    "3Signs",
    "3Symbols",
    "3Symbols2",
    "3Stars",
    "3Triangles",
]


def apply_three_icon_rule(
    file_path: Path,
    sheet_name: str,
    cell_range: str,
    icon_style: str = "3Symbols2",
    reverse: bool = False,
    show_values: bool = True,
    icons_only: bool = False,
    coerce_numbers: bool = False,
    non_numeric_to_zero: bool = False,
    skip_hide_number_format: bool = False,
) -> None:
    """
    Apply a 3-icon conditional formatting rule to the given range.

    The rule splits values into three buckets around zero:
    - top icon: > 0
    - middle icon: = 0
    - bottom icon: < 0

    This uses an IconSet with numeric thresholds set to [0, 0].
    """

    if icon_style not in ALLOWED_ICON_STYLES:
        raise ValueError(
            f"icon_style must be one of: {', '.join(ALLOWED_ICON_STYLES)}"
        )

    if not file_path.exists():
        raise FileNotFoundError(f"Workbook not found: {file_path}")

    wb = load_workbook(filename=str(file_path))
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}"
        )

    ws = wb[sheet_name]

    # Coerce values in the target range to numbers before applying icons
    if coerce_numbers:
        _coerce_range_to_numbers(ws, cell_range, non_numeric_to_zero)

    # Build rule using low-level API to ensure persistence in Excel CF manager
    # Thresholds: first is strictly > 0 (gte=False), second is >= 0 (gte=True)
    # This yields: top icon for >0, middle for =0, bottom for <0
    cfvo = [
        FormatObject(type="num", val=0, gte=False),
        FormatObject(type="num", val=0, gte=True),
    ]
    iconset = IconSet(
        iconSet=icon_style,
        cfvo=cfvo,
        showValue=show_values,
        percent=False,
        reverse=reverse,
    )
    rule = Rule(type="iconSet", iconSet=iconset)

    ws.conditional_formatting.add(cell_range, rule)

    # Reliability fallback: hide numbers via number format when icons_only
    # Number format ';;;' hides numeric display entirely in Excel.
    if icons_only and not skip_hide_number_format:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.number_format = ';;;'

    wb.save(str(file_path))


def apply_number_format_icons(
    file_path: Path,
    sheet_name: str,
    cell_range: str,
    number_pattern: str = "0",
    icons_only: bool = False,
    pos_symbol: str = "✓↑",
    zero_symbol: str = "⚑",
    neg_symbol: str = "✗",
    pos_color: str = "Color4",   # Bright Green
    zero_color: str = "Color46", # Orange
    neg_color: str = "Color3",   # Red
) -> None:
    """
    Apply a custom number format that displays icon-like symbols by sign:

    - positive (>0): orange flag (via [Color46] and Unicode flag symbol)
    - zero (=0): green check mark
    - negative (<0): red X mark

    Number format used: [Color46]"⚑ "0;[Red]"✗ "0;[Green]"✓ "0
    """

    if not file_path.exists():
        raise FileNotFoundError(f"Workbook not found: {file_path}")

    wb = load_workbook(filename=str(file_path))
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}"
        )

    ws = wb[sheet_name]

    # Build a three-section number format: positive;negative;zero
    # Users can change the numeric pattern (e.g., "0.00" or "#,##0").
    def color_token(c: str) -> str:
        # Wrap color name/index in brackets as Excel expects, e.g., [Green], [Color46]
        c = c.strip()
        if c.startswith("[") and c.endswith("]"):
            return c
        return f"[{c}]"

    pos_c = color_token(pos_color)
    zero_c = color_token(zero_color)
    neg_c = color_token(neg_color)

    if icons_only:
        # Show only the icons (replace numbers completely)
        pos = f"{pos_c} \"{pos_symbol}\""
        neg = f"{neg_c} \"{neg_symbol}\""
        zer = f"{zero_c} \"{zero_symbol}\""
    else:
        pos = f"{pos_c} \"{pos_symbol} \"{number_pattern}"
        neg = f"{neg_c} \"{neg_symbol} \"{number_pattern}"
        zer = f"{zero_c} \"{zero_symbol} \"{number_pattern}"
    fmt = f"{pos};{neg};{zer}"

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.number_format = fmt

    wb.save(str(file_path))


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description=(
            "Apply 3-icon conditional formatting (>0, =0, <0) to a sheet range in an Excel file."
        )
    )
    parser.add_argument(
        "--mode",
        choices=["iconset", "number-format"],
        default="iconset",
        help=(
            "How to display icons: 'iconset' uses Excel Icon Sets; 'number-format' uses a custom number format with Unicode symbols."
        ),
    )
    parser.add_argument(
        "--file",
        required=True,
        help="Path to the .xlsx workbook",
    )
    parser.add_argument(
        "--sheet",
        required=True,
        help="Target worksheet name",
    )
    parser.add_argument(
        "--range",
        dest="cell_range",
        required=True,
        help="A1-style range, e.g. A2:D200",
    )
    parser.add_argument(
        "--icon-style",
        default="3Symbols2",
        choices=ALLOWED_ICON_STYLES,
        help="3-icon set style to use",
    )
    parser.add_argument(
        "--reverse",
        action="store_true",
        help="Reverse icon order (e.g., green->red)",
    )
    parser.add_argument(
        "--icons-only",
        action="store_true",
        help="Show icons without underlying numeric values",
    )
    parser.add_argument(
        "--hide-values",
        action="store_true",
        help="Hide numeric values (alias of --icons-only)",
    )
    parser.add_argument(
        "--coerce-numbers",
        action="store_true",
        help="Convert values in the given range to numbers before formatting.",
    )
    parser.add_argument(
        "--coerce-non-numeric-zero",
        action="store_true",
        help="When coercing numbers, set non-numeric cells to 0.",
    )
    parser.add_argument(
        "--nf-pattern",
        default="0",
        help="Numeric pattern for number-format mode (e.g., '0.00', '#,##0').",
    )
    parser.add_argument(
        "--nf-icons-only",
        action="store_true",
        help="In number-format mode, hide numeric values and show icons only.",
    )
    parser.add_argument(
        "--nf-add-placeholder-iconset",
        action="store_true",
        help=(
            "In number-format mode, also add a built-in IconSet rule (visible in Excel CF Manager) before applying custom number-format icons."
        ),
    )
    parser.add_argument("--nf-pos-symbol", default="✓↑", help="Symbol for positive (>0) values in number-format mode.")
    parser.add_argument("--nf-zero-symbol", default="⚑", help="Symbol for zero (=0) values in number-format mode.")
    parser.add_argument("--nf-neg-symbol", default="✗", help="Symbol for negative (<0) values in number-format mode.")
    parser.add_argument("--nf-pos-color", default="Color4", help="Color name or index (e.g., 'Green' or 'Color46') for positive values.")
    parser.add_argument("--nf-zero-color", default="Color46", help="Color name or index for zero values.")
    parser.add_argument("--nf-neg-color", default="Color3", help="Color name or index for negative values.")

    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)

    file_path = Path(args.file)
    if args.mode == "iconset":
        show_values = not (args.icons_only or args.hide_values)
        apply_three_icon_rule(
            file_path=file_path,
            sheet_name=args.sheet,
            cell_range=args.cell_range,
            icon_style=args.icon_style,
            reverse=args.reverse,
            show_values=show_values,
            icons_only=(args.icons_only or args.hide_values),
            coerce_numbers=args.coerce_numbers,
            non_numeric_to_zero=args.coerce_non_numeric_zero,
            skip_hide_number_format=False,
        )
        print(
            f"Applied {args.icon_style} icon set to {args.sheet}!{args.cell_range} in {file_path}"
        )
    else:
        # Optional hybrid: add a placeholder IconSet rule for visibility, then apply custom number-format icons for exact visuals
        if args.nf_add_placeholder_iconset:
            apply_three_icon_rule(
                file_path=file_path,
                sheet_name=args.sheet,
                cell_range=args.cell_range,
                icon_style=args.icon_style,
                reverse=args.reverse,
                show_values=False,
                icons_only=True,
                coerce_numbers=args.coerce_numbers,
                non_numeric_to_zero=args.coerce_non_numeric_zero,
                skip_hide_number_format=True,
            )
        apply_number_format_icons(
            file_path=file_path,
            sheet_name=args.sheet,
            cell_range=args.cell_range,
            number_pattern=args.nf_pattern,
            icons_only=(args.nf_icons_only or args.icons_only or args.hide_values),
            pos_symbol=args.nf_pos_symbol,
            zero_symbol=args.nf_zero_symbol,
            neg_symbol=args.nf_neg_symbol,
            pos_color=args.nf_pos_color,
            zero_color=args.nf_zero_color,
            neg_color=args.nf_neg_color,
        )
        print(
            (
                "Applied number-format icons (orange flag, green check, red X)"
                + (" with placeholder IconSet rule" if args.nf_add_placeholder_iconset else "")
                + f" to {args.sheet}!{args.cell_range} in {file_path}"
            )
        )


if __name__ == "__main__":
    sys.exit(main())
